import json
import re
import shutil
from datetime import datetime, date, time
from pathlib import Path

from openpyxl import load_workbook


# =========================================================
# AYARLAR
# =========================================================

SOURCE = Path("data.xlsx")
OUTPUT = Path("data")

DEFAULT_SHEET = "TICKET"


# =========================================================
# DOSYA ADI
# =========================================================

def safe_filename(name):
    """
    Sheet adını güvenli dosya adına çevirir.
    """
    name = str(name)

    name = re.sub(
        r'[<>:"/\\|?*]',
        "_",
        name
    )

    name = name.strip(" .")

    return name or "sheet"


# =========================================================
# EXCEL DEĞERİNİ JSON'A ÇEVİR
# =========================================================

def clean_value(value):
    """
    Excel'den gelen değeri JSON için temizler.

    ÖNEMLİ:
    datetime -> SADECE TARİH
    Saat kısmı kesinlikle yazılmaz.
    Örn:
    29.08.2026
    """

    if value is None:
        return ""


    # -----------------------------------------------------
    # datetime
    # -----------------------------------------------------

    if isinstance(value, datetime):

        return value.strftime(
            "%d.%m.%Y"
        )


    # -----------------------------------------------------
    # date
    # -----------------------------------------------------

    if isinstance(value, date):

        return value.strftime(
            "%d.%m.%Y"
        )


    # -----------------------------------------------------
    # time
    # -----------------------------------------------------

    if isinstance(value, time):

        # Kullanıcı saat istemediği için
        # saat hücrelerini boş bırakıyoruz.
        return ""


    # -----------------------------------------------------
    # Diğer JSON uyumlu değerler
    # -----------------------------------------------------

    if isinstance(
        value,
        (
            str,
            int,
            float,
            bool
        )
    ):

        return value


    # -----------------------------------------------------
    # Bilinmeyen tip
    # -----------------------------------------------------

    return str(value)


# =========================================================
# ANA İŞLEM
# =========================================================

def main():

    print()
    print("=" * 60)
    print("PS26 EXCEL -> JSON DONUSTURUCU")
    print("=" * 60)
    print()


    # -----------------------------------------------------
    # Excel kontrolü
    # -----------------------------------------------------

    if not SOURCE.exists():

        print(
            "HATA: data.xlsx bulunamadı!"
        )

        print(
            f"Aranan dosya: {SOURCE.resolve()}"
        )

        print()

        input(
            "Kapatmak için Enter'a basın..."
        )

        return


    # -----------------------------------------------------
    # Eski data klasörünü tamamen sil
    # -----------------------------------------------------

    if OUTPUT.exists():

        print(
            "Eski data klasörü temizleniyor..."
        )

        shutil.rmtree(
            OUTPUT
        )


    OUTPUT.mkdir(
        parents=True,
        exist_ok=True
    )


    # -----------------------------------------------------
    # Excel'i read-only aç
    # -----------------------------------------------------

    print(
        "Excel okunuyor..."
    )

    print(
        f"Kaynak: {SOURCE.resolve()}"
    )

    print()


    workbook = load_workbook(
        SOURCE,
        read_only=True,
        data_only=True
    )


    sheet_index = []


    # -----------------------------------------------------
    # TÜM SHEETLER
    # -----------------------------------------------------

    for sheet in workbook.worksheets:

        sheet_name = sheet.title

        print(
            f"İşleniyor: {sheet_name}"
        )


        rows_iterator = sheet.iter_rows(
            values_only=True
        )


        # -------------------------------------------------
        # BAŞLIK SATIRI
        # -------------------------------------------------

        try:

            header_values = next(
                rows_iterator
            )

        except StopIteration:

            header_values = []


        headers = [
            clean_value(value)
            for value in header_values
        ]


        # -------------------------------------------------
        # VERİ SATIRLARI
        # -------------------------------------------------

        data_rows = []

        for row in rows_iterator:

            cleaned = [
                clean_value(value)
                for value in row
            ]


            # -------------------------------------------------
            # Tamamen boş satırı at
            # -------------------------------------------------

            if not any(
                value != ""
                for value in cleaned
            ):

                continue


            # -------------------------------------------------
            # Başlık sayısına göre sütunları düzelt
            # -------------------------------------------------

            if len(cleaned) < len(headers):

                cleaned.extend(
                    [""] *
                    (
                        len(headers)
                        -
                        len(cleaned)
                    )
                )


            elif len(cleaned) > len(headers):

                cleaned = cleaned[
                    :len(headers)
                ]


            data_rows.append(
                cleaned
            )


        # -------------------------------------------------
        # SHEET JSON
        # -------------------------------------------------

        sheet_data = {
            "name": sheet_name,
            "headers": headers,
            "rows": data_rows
        }


        filename = (
            safe_filename(sheet_name)
            +
            ".json"
        )


        output_file = (
            OUTPUT
            /
            filename
        )


        # -------------------------------------------------
        # JSON YAZ
        # -------------------------------------------------

        with open(
            output_file,
            "w",
            encoding="utf-8"
        ) as file:

            json.dump(
                sheet_data,
                file,
                ensure_ascii=False,
                separators=(",", ":")
            )


        # -------------------------------------------------
        # INDEX BİLGİSİ
        # -------------------------------------------------

        sheet_index.append({
            "name": sheet_name,
            "file": filename,
            "rows": len(data_rows),
            "columns": len(headers)
        })


        print(
            f"  -> {filename}"
        )

        print(
            f"  -> {len(data_rows):,} satır"
        )

        print(
            f"  -> {len(headers):,} sütun"
        )

        print()


    # -----------------------------------------------------
    # VARSAYILAN SHEET
    # -----------------------------------------------------

    default_sheet = ""

    for item in sheet_index:

        if item["name"] == DEFAULT_SHEET:

            default_sheet = DEFAULT_SHEET

            break


    # TICKET yoksa ilk sheet
    if not default_sheet and sheet_index:
        default_sheet = sheet_index[0]["name"]


    # -----------------------------------------------------
    # INDEX.JSON
    # -----------------------------------------------------

    index_data = {
        "defaultSheet": default_sheet,
        "sheets": sheet_index
    }


    with open(
        OUTPUT / "index.json",
        "w",
        encoding="utf-8"
    ) as file:

        json.dump(
            index_data,
            file,
            ensure_ascii=False,
            separators=(",", ":")
        )


    workbook.close()


    # -----------------------------------------------------
    # SONUÇ
    # -----------------------------------------------------

    print()
    print("=" * 60)
    print("TAMAMLANDI")
    print("=" * 60)
    print()

    print(
        f"{len(sheet_index)} sheet oluşturuldu."
    )

    print()

    print(
        "Varsayılan sheet: "
        +
        default_sheet
    )

    print()

    print(
        "Oluşan sheetler:"
    )

    for item in sheet_index:

        print(
            f"  - {item['name']}: "
            f"{item['rows']:,} satır / "
            f"{item['columns']:,} sütun"
        )

    print()

    print(
        f"Çıktı klasörü: "
        f"{OUTPUT.resolve()}"
    )

    print()


if __name__ == "__main__":

    main()